from openpyxl import load_workbook

f_out = open('./overzicht downloads.txt','w+')

wb = load_workbook('./MOOC ROS (Robot Operating System).xlsx')

sheet = wb['Blad1']

for i in range(6,4000):
    print('lec: {} title: {}'.format(sheet.cell(row=i,column=1).value , sheet.cell(row=i,column=2).value))
    print(' ')
    print('<iframe class="dframe"')
    print('       data-downloadid="{}"'.format( sheet.cell(row=i,column=9).value))
    print('       data-youtubeid="{}"></iframe>'.format ( sheet.cell(row=i,column=8).value))
    print('<script src="https://delftxdownloads.tudelft.nl/_tools/dl/js/builder.js"></script>')
    print(' ')
    print(' ')
    
    if  sheet.cell(row=i,column=9).value != None :
        print('leeg')
        f_out.write('lec: {} title: {}\n'.format(sheet.cell(row=i,column=1).value , sheet.cell(row=i,column=2).value))
        f_out.write(' \n')
        f_out.write('<iframe class="dframe"\n')
        f_out.write('       data-downloadid="{}"\n'.format( sheet.cell(row=i,column=9).value))
        f_out.write('       data-youtubeid="{}"></iframe>\n'.format ( sheet.cell(row=i,column=8).value))
        f_out.write('<script src="https://delftxdownloads.tudelft.nl/_tools/dl/js/builder.js"></script>\n')
        f_out.write(' \n')
        f_out.write(' \n')
        f_out.write('-----------------------------------------------------------\n')
        f_out.write(' \n')
